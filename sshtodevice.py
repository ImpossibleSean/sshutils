import openpyxl
import sys
import getpass
import pprint
from netmiko import SSHDetect
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko.ssh_exception import NetMikoAuthenticationException


def excelimportcol(filename, sheetname='Sheet1'):
    # Creates a key:value pair dictionary based on the key being the first item in the column
    # and adds all dictionaries to a list.
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[sheetname]
    excellist = []
    for row in range(2, worksheet.max_row + 1):
        config = {}
        for col in range(1, worksheet.max_column + 1):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            config[str(key)] = str(value)
        excellist.append(config)
    return excellist


def kwargsfromdict(device):
    if 'device_type' not in device or \
            device['device_type'] == '' or \
            device['device_type'] is None:
        device_type = 'autodetect'
    else:
        device_type = device['device_type']
    connect_kwargs = {'device_type': device_type,
                      'host': device['IP'],
                      'username': device['username'],
                      'password': device['password']}
    if device_type == 'autodetect':
        connect_kwargs['device_type'] = autodetect(connect_kwargs)

    return connect_kwargs


def autodetect(connect_kwargs):
    try:
        net_connect = SSHDetect(**connect_kwargs)
        print(net_connect.autodetect())
        return net_connect.autodetect()

    except NetMikoTimeoutException:
        print('The IP address: {} did not respond, it might be unreachable'.format(connect_kwargs['host']))
    except NetMikoAuthenticationException:
        print('Authentication failed for IP address: {}'.format(connect_kwargs['host']))


def get_connection(connect_kwargs):
    try:
        return ConnectHandler(**connect_kwargs)

    except NetMikoTimeoutException:
        print('The IP address: {} did not respond, it might be unreachable'.format(connect_kwargs['host']))
    except NetMikoAuthenticationException:
        print('Authentication failed for IP address: {}'.format(connect_kwargs['host']))


def sendcommands(net_connect, commandset, name):
    print('Connecting to ' + name)
    for command in commandset:
        output = net_connect.send_command(command)
        print(output)



def removefalsecommits(devicelist):
    for devicedict in devicelist:
        if devicedict['Commit'] != 'True':
            pprint.pprint('Removing element ')
            pprint.pprint(devicedict)
            devicelist.remove(devicedict)


def listdictappend(listofdicts, dict2):
    for dict1 in listofdicts:
        dict1.update( dict2 )

def readcommands(filename):
    configlines = []
    f = open(filename, "r")
    for line in f.readlines():
        configlines.append(line)
    return configlines


# username = sys.argv[1]
# password = sys.argv[2]
username = 'nolooking'
password = getpass.getpass()




devices = excelimportcol('devices.xlsx')
removefalsecommits(devices)
deviceappend = {'username': username, 'password': password}
listdictappend(devices, deviceappend)
commandset = readcommands('config.txt')
pprint.pprint(devices)
for device in devices:
    devicekwargs = kwargsfromdict(device)
    connection = get_connection(devicekwargs)
    sendcommands(connection, commandset, device['Hostname'])